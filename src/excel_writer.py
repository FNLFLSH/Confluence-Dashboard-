import json
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .utils import standardize_fields, add_quarter_column
from collections import defaultdict

# Pre-compile regex for quarter sorting
QUARTER_SORT_PATTERN = re.compile(r'(\d{4}) Q(\d)')

def get_quarter_sort_key(q_string):
    """Generate a sort key for quarter strings (e.g., '2024 Q1 (Jan–Mar)')."""
    match = QUARTER_SORT_PATTERN.search(q_string)
    if match:
        year = int(match.group(1))
        quarter_num = int(match.group(2))
        # Sort by year first (descending - most recent first), then by quarter number
        return (-year, quarter_num)
    # Place "Unknown" or malformed quarters at the end
    return (9999, 5)

def analyze_module_changes(data):
    """
    Analyze modules that have undergone multiple changes within quarters and years.
    Returns data for modules with 2+ changes per quarter and 3+ changes per year.
    """
    # Group by module name and quarter/year
    module_quarter_changes = defaultdict(lambda: defaultdict(list))
    module_year_changes = defaultdict(lambda: defaultdict(list))
    
    for item in data:
        module_name = item.get('ModuleName', '').strip()
        if not module_name:  # Skip items without module names
            continue
            
        quarter = item.get('Quarter', '')
        category = item.get('Category', 'Other')
        date = item.get('Date', '')
        
        # Extract year from quarter or date
        year = None
        if quarter:
            year_match = re.search(r'(\d{4})', quarter)
            if year_match:
                year = year_match.group(1)
        elif date:
            try:
                if isinstance(date, str):
                    if 'T' in date:
                        date = date.split('T')[0]
                    year = datetime.strptime(date, '%Y-%m-%d').year
                elif isinstance(date, datetime):
                    year = date.year
            except (ValueError, TypeError):
                pass
        
        if quarter:
            module_quarter_changes[module_name][quarter].append({
                'title': item.get('Title', ''),
                'category': category,
                'date': date
            })
        
        if year:
            module_year_changes[module_name][str(year)].append({
                'title': item.get('Title', ''),
                'category': category,
                'date': date
            })
    
    # Find modules with multiple changes
    frequent_quarter_changes = []
    frequent_year_changes = []
    
    for module_name, quarters in module_quarter_changes.items():
        for quarter, changes in quarters.items():
            if len(changes) >= 2:  # 2 or more changes in a quarter
                frequent_quarter_changes.append({
                    'module_name': module_name,
                    'period': quarter,
                    'change_count': len(changes),
                    'changes': changes
                })
    
    for module_name, years in module_year_changes.items():
        for year, changes in years.items():
            if len(changes) >= 3:  # 3 or more changes in a year
                frequent_year_changes.append({
                    'module_name': module_name,
                    'period': year,
                    'change_count': len(changes),
                    'changes': changes
                })
    
    # Sort by change count (descending), then by module name
    frequent_quarter_changes.sort(key=lambda x: (-x['change_count'], x['module_name']))
    frequent_year_changes.sort(key=lambda x: (-x['change_count'], x['module_name']))
    
    return frequent_quarter_changes, frequent_year_changes

def export_grouped_by_quarter(data, filename):
    """
    Export release reports to a styled Excel sheet grouped by quarter.
    Optimized for speed with reduced operations and better data structures.
    """
    if isinstance(data, str):
        data = json.loads(data)
    
    # Standardize the data
    data = standardize_fields(data)
    data = add_quarter_column(data)

    wb = Workbook()
    ws = wb.active
    ws.title = "Release Notes"

    # --- Pre-define styles for better performance ---
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0033A0", end_color="0033A0", fill_type="solid")
    quarter_header_font = Font(bold=True, size=14, color="0033A0", underline="single")
    main_title_font = Font(bold=True, size=18, color="0033A0")
    total_font = Font(bold=True)
    align_wrap_top = Alignment(wrap_text=True, vertical="top", horizontal="left")
    align_center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    # --- Main Title ---
    ws.merge_cells('A1:E1')
    main_title_cell = ws['A1']
    main_title_cell.value = "Release Reports by Quarter"
    main_title_cell.font = main_title_font
    main_title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30
    
    row = 3
    
    # Get unique quarters and sort them efficiently
    quarter_data = {}
    for item in data:
        quarter = item.get('Quarter')
        if quarter:
            if quarter not in quarter_data:
                quarter_data[quarter] = []
            quarter_data[quarter].append(item)
    
    quarters = sorted(quarter_data.keys(), key=get_quarter_sort_key)
    
    summary = {}

    for q in quarters:
        # Quarter header
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        q_header_cell = ws.cell(row=row, column=1, value=q)
        q_header_cell.font = quarter_header_font
        row += 1

        # Table headers
        headers = ["Report", "Details", "Module Name", "Category", "Date"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = thin_border
        row += 1

        # Process quarter data - filter out new module releases
        quarter_items = [item for item in quarter_data[q] if not (
            item.get("NewRelease", False) or 
            "new release" in (item.get("Body", "") or "").lower() or
            "new release" in (item.get("Title", "") or "").lower()
        )]
        counts = {}
        
        for report in quarter_items:
            # Write data efficiently
            ws.cell(row=row, column=1, value=report.get("Title"))
            ws.cell(row=row, column=2, value=report.get("Body"))
            ws.cell(row=row, column=3, value=report.get("ModuleName", ""))
            ws.cell(row=row, column=4, value=report.get("Category"))
            
            # Optimized date handling
            date_val = report.get("Date")
            if isinstance(date_val, str):
                try:
                    if date_val.endswith('Z'):
                        date_val = date_val[:-1]
                    if 'T' in date_val:
                        date_val = date_val.split('T')[0]
                    date_val = datetime.fromisoformat(date_val)
                except (ValueError, TypeError):
                    try:
                        date_val = datetime.strptime(date_val, '%Y-%m-%d')
                    except (ValueError, TypeError):
                        pass
            
            # Format date for display
            if isinstance(date_val, datetime):
                date_cell = ws.cell(row=row, column=5, value=date_val.strftime('%Y-%m-%d'))
            else:
                date_cell = ws.cell(row=row, column=5, value=str(date_val) if date_val else "")
            
            # Apply styling efficiently
            for col in range(1, 6):
                cell = ws.cell(row=row, column=col)
                cell.alignment = align_wrap_top
                cell.border = thin_border
            row += 1

            # Count categories
            cat = report.get('Category', 'Other')
            counts[cat] = counts.get(cat, 0) + 1
        
        summary[q] = counts
        row += 1

    # --- Summary Section ---
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    summary_title_cell = ws.cell(row=row, column=1, value="📊 Summary")
    summary_title_cell.font = quarter_header_font
    row += 1
    
    summary_headers = ["Quarter", "Bug Fix", "Enhancement", "New Feature", "Other"]
    all_cats = ["Bug Fix", "Enhancement", "New Feature", "Other"]

    # Summary headers
    for col_idx, header in enumerate(summary_headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border
    row += 1

    grand_total = {cat: 0 for cat in all_cats}
    
    # Group quarters by year for yearly totals
    yearly_data = {}
    for quarter in quarters:
        year_match = re.search(r'(\d{4})', quarter)
        if year_match:
            year = year_match.group(1)
            if year not in yearly_data:
                yearly_data[year] = []
            yearly_data[year].append(quarter)
    
    # Summary data with yearly totals
    for quarter in quarters:
        counts = summary.get(quarter, {})
        q_cell = ws.cell(row=row, column=1, value=quarter)
        q_cell.border = thin_border
        
        for idx, cat in enumerate(all_cats, 2):
            count = counts.get(cat, 0)
            cell = ws.cell(row=row, column=idx, value=count)
            cell.alignment = align_center
            cell.border = thin_border
            grand_total[cat] += count
        row += 1
        
        # Check if this is the last quarter of a year and add yearly total
        year_match = re.search(r'(\d{4})', quarter)
        if year_match:
            year = year_match.group(1)
            year_quarters = yearly_data.get(year, [])
            if quarter == year_quarters[-1]:  # Last quarter of this year
                # Calculate yearly total for this year
                yearly_total = {cat: 0 for cat in all_cats}
                for q in year_quarters:
                    q_counts = summary.get(q, {})
                    for cat in all_cats:
                        yearly_total[cat] += q_counts.get(cat, 0)
                
                # Add yearly total row
                yearly_total_cell = ws.cell(row=row, column=1, value=f"{year} Total")
                yearly_total_cell.font = total_font
                yearly_total_cell.border = thin_border
                
                for idx, cat in enumerate(all_cats, 2):
                    cell = ws.cell(row=row, column=idx, value=yearly_total[cat])
                    cell.font = total_font
                    cell.alignment = align_center
                    cell.border = thin_border
                row += 1

    # --- Auto-size Columns ---
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    
    # Optimize row heights
    for i in range(1, row + 1):
        ws.row_dimensions[i].height = max(15, ws.row_dimensions[i].height or 15)

    # --- Create New Releases Worksheet ---
    create_new_releases_worksheet(wb, data)

    # --- Create Frequent Changes Worksheet ---
    create_frequent_changes_worksheet(wb, data)

    # Save the workbook
    wb.save(filename)
    print(f"Excel report generated: {filename}")
    print(f"Total releases processed: {len(data)}")
    print(f"Quarters found: {len(quarters)}")
    print(f"Categories: {list(set(item['Category'] for item in data))}")

def export_frequent_changes_report(data, filename):
    """
    Export a separate report showing modules with frequent changes.
    Creates a dedicated Excel file for frequent module changes analysis.
    """
    if isinstance(data, str):
        data = json.loads(data)
    
    # Standardize the data
    data = standardize_fields(data)
    data = add_quarter_column(data)
    
    # Analyze frequent changes
    print("Analyzing modules with frequent changes...")
    frequent_quarter_changes, frequent_year_changes = analyze_module_changes(data)
    
    # Create new workbook for frequent changes
    wb = Workbook()
    ws = wb.active
    ws.title = "Frequent Module Changes"
    
    # --- Pre-define styles for better performance ---
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0033A0", end_color="0033A0", fill_type="solid")
    quarter_header_font = Font(bold=True, size=14, color="0033A0", underline="single")
    main_title_font = Font(bold=True, size=18, color="0033A0")
    total_font = Font(bold=True)
    align_wrap_top = Alignment(wrap_text=True, vertical="top", horizontal="left")
    align_center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    # --- Main Title ---
    ws.merge_cells('A1:E1')
    main_title_cell = ws['A1']
    main_title_cell.value = "Modules with Frequent Changes - Analysis Report"
    main_title_cell.font = main_title_font
    main_title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30
    
    row = 3
    
    # --- Description Section ---
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    desc_title = ws.cell(row=row, column=1, value="📋 Analysis Logic & Purpose")
    desc_title.font = Font(bold=True, size=12, color="0033A0")
    row += 1
    
    # Add description text
    description_text = [
        "This report identifies modules that have undergone frequent changes, helping to:",
        "• Identify modules requiring special attention or refactoring",
        "• Track development velocity and maintenance patterns", 
        "• Highlight potential technical debt or stability issues",
        "",
        "Quarterly Analysis: Modules with 2+ changes in a single quarter",
        "Yearly Analysis: Modules with 3+ changes in a single year",
        "",
        "Note: Each row represents a module that exceeded the threshold in that specific time period."
    ]
    
    for desc_line in description_text:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        desc_cell = ws.cell(row=row, column=1, value=desc_line)
        desc_cell.font = Font(size=10)
        desc_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        row += 1
    
    row += 2  # Add spacing
    
    # --- Quarterly Changes Section ---
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    q_frequent_title = ws.cell(row=row, column=1, value="📈 Quarterly Analysis: Modules with 2+ Changes per Quarter")
    q_frequent_title.font = quarter_header_font
    row += 1
    
    if frequent_quarter_changes:
        # Group quarterly changes by quarter for better organization
        quarter_groups = {}
        for item in frequent_quarter_changes:
            quarter = item['period']
            if quarter not in quarter_groups:
                quarter_groups[quarter] = []
            quarter_groups[quarter].append(item)
        
        # Sort quarters
        sorted_quarters = sorted(quarter_groups.keys(), key=get_quarter_sort_key)
        
        for quarter in sorted_quarters:
            # Quarter sub-header
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            quarter_subtitle = ws.cell(row=row, column=1, value=f"Quarter: {quarter}")
            quarter_subtitle.font = Font(bold=True, size=11, color="2E5C8A")
            quarter_subtitle.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
            row += 1
            
            # Headers for this quarter
            q_headers = ["Module Name", "Change Count", "Changes Summary", "Change Details"]
            for col_idx, header in enumerate(q_headers, 1):
                cell = ws.cell(row=row, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = align_center
                cell.border = thin_border
            row += 1
            
            # Data for this quarter
            for item in quarter_groups[quarter]:
                ws.cell(row=row, column=1, value=item['module_name'])
                ws.cell(row=row, column=2, value=item['change_count'])
                
                # Create summary text
                categories = {}
                for change in item['changes']:
                    cat = change['category']
                    categories[cat] = categories.get(cat, 0) + 1
                
                summary_parts = []
                for cat, count in categories.items():
                    summary_parts.append(f"{cat}: {count}")
                summary_text = ", ".join(summary_parts)
                
                ws.cell(row=row, column=3, value=summary_text)
                
                # Format detailed changes
                changes_text = []
                for i, change in enumerate(item['changes'], 1):
                    change_str = f"{i}. {change['category']}: {change['title']}"
                    changes_text.append(change_str)
                
                changes_cell = ws.cell(row=row, column=4, value="\n".join(changes_text))
                changes_cell.alignment = align_wrap_top
                
                # Apply styling
                for col in range(1, 5):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border
                row += 1
            
            row += 1  # Add spacing between quarters
    else:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        no_q_data = ws.cell(row=row, column=1, value="✅ No modules found with 2+ changes per quarter")
        no_q_data.font = Font(italic=True, color="008000")
        row += 1
    
    row += 2  # Add spacing before yearly section
    
    # --- Yearly Changes Section ---
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    y_frequent_title = ws.cell(row=row, column=1, value="📊 Yearly Analysis: Modules with 3+ Changes per Year")
    y_frequent_title.font = quarter_header_font
    row += 1
    
    if frequent_year_changes:
        # Headers for yearly changes
        y_headers = ["Module Name", "Year", "Total Changes", "Quarterly Breakdown", "Change Details"]
        for col_idx, header in enumerate(y_headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = thin_border
        row += 1
        
        # Data for yearly changes
        for item in frequent_year_changes:
            ws.cell(row=row, column=1, value=item['module_name'])
            ws.cell(row=row, column=2, value=item['period'])
            ws.cell(row=row, column=3, value=item['change_count'])
            
            # Create quarterly breakdown
            quarter_counts = {}
            for change in item['changes']:
                # Extract quarter from date if possible
                date_str = str(change['date'])
                try:
                    if 'T' in date_str:
                        date_str = date_str.split('T')[0]
                    date_obj = datetime.strptime(date_str, '%Y-%m-%d')
                    quarter = f"Q{(date_obj.month-1)//3 + 1}"
                    quarter_counts[quarter] = quarter_counts.get(quarter, 0) + 1
                except:
                    quarter_counts['Unknown'] = quarter_counts.get('Unknown', 0) + 1
            
            breakdown_text = ", ".join([f"{q}: {c}" for q, c in sorted(quarter_counts.items())])
            ws.cell(row=row, column=4, value=breakdown_text)
            
            # Format detailed changes
            changes_text = []
            for i, change in enumerate(item['changes'], 1):
                change_str = f"{i}. {change['category']}: {change['title']}"
                changes_text.append(change_str)
            
            changes_cell = ws.cell(row=row, column=5, value="\n".join(changes_text))
            changes_cell.alignment = align_wrap_top
            
            # Apply styling
            for col in range(1, 6):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
            row += 1
    else:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        no_y_data = ws.cell(row=row, column=1, value="✅ No modules found with 3+ changes per year")
        no_y_data.font = Font(italic=True, color="008000")
        row += 1
    
    # --- Auto-size columns ---
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    
    # Optimize row heights
    for i in range(1, row + 1):
        ws.row_dimensions[i].height = max(15, ws.row_dimensions[i].height or 15)
    
    # Save the workbook
    wb.save(filename)
    print(f"Frequent changes report generated: {filename}")
    print(f"Frequent changes analysis: {len(frequent_quarter_changes)} quarterly, {len(frequent_year_changes)} yearly")

def create_new_releases_worksheet(wb, data):
    """
    Create a new releases worksheet in the existing workbook.
    """
    ws = wb.create_sheet("New Releases")
    
    # Filter new releases
    new_releases = [item for item in data if (
        item.get("NewRelease", False) or 
        "new release" in (item.get("Body", "") or "").lower() or
        "new release" in (item.get("Title", "") or "").lower()
    )]
    
    # Sort by date
    new_releases.sort(key=lambda x: x.get("Date", ""))
    
    # --- Pre-define styles ---
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    title_font = Font(bold=True, size=16, color="0033A0")
    align_center = Alignment(horizontal="center", vertical="center")
    align_wrap = Alignment(wrap_text=True, vertical="top")
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    # --- Title ---
    ws.merge_cells('A1:C1')
    title_cell = ws['A1']
    title_cell.value = "New Module Releases"
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30
    
    row = 3
    
    # --- Headers ---
    headers = ["Date", "Module Name", "Details"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border
    row += 1
    
    # --- Data ---
    for release in new_releases:
        # Format date
        date_val = release.get("Date")
        if isinstance(date_val, str):
            try:
                if date_val.endswith('Z'):
                    date_val = date_val[:-1]
                if 'T' in date_val:
                    date_val = date_val.split('T')[0]
                date_val = datetime.fromisoformat(date_val)
            except (ValueError, TypeError):
                try:
                    date_val = datetime.strptime(date_val, '%Y-%m-%d')
                except (ValueError, TypeError):
                    pass
        
        if isinstance(date_val, datetime):
            date_val = date_val.strftime('%Y-%m-%d')
        else:
            date_val = str(date_val) if date_val else ""
        
        # Add row data
        ws.cell(row=row, column=1, value=date_val)
        ws.cell(row=row, column=2, value=release.get("ModuleName", ""))
        details_cell = ws.cell(row=row, column=3, value=release.get("Body", ""))
        details_cell.alignment = align_wrap
        
        # Apply borders
        for col in range(1, 4):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
        
        row += 1
    
    # --- Column widths ---
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 80
    
    print(f"New releases worksheet created with {len(new_releases)} releases")

def create_frequent_changes_worksheet(wb, data):
    """
    Create a frequent changes worksheet in the existing workbook.
    """
    ws = wb.create_sheet("Frequent Changes")
    
    # Analyze frequent changes
    print("Analyzing modules with frequent changes...")
    frequent_quarter_changes, frequent_year_changes = analyze_module_changes(data)
    
    # --- Pre-define styles ---
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0033A0", end_color="0033A0", fill_type="solid")
    title_font = Font(bold=True, size=16, color="0033A0")
    subtitle_font = Font(bold=True, size=12, color="0033A0")
    align_center = Alignment(horizontal="center", vertical="center")
    align_wrap = Alignment(wrap_text=True, vertical="top")
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    # --- Title ---
    ws.merge_cells('A1:D1')
    title_cell = ws['A1']
    title_cell.value = "Modules with Frequent Changes"
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30
    
    row = 3
    
    # --- Description ---
    ws.merge_cells(f'A{row}:D{row}')
    desc_cell = ws.cell(row=row, column=1, value="Analysis of modules with 2+ changes per quarter or 3+ changes per year")
    desc_cell.font = Font(italic=True)
    row += 2
    
    # --- Quarterly Changes Section ---
    ws.merge_cells(f'A{row}:D{row}')
    q_title = ws.cell(row=row, column=1, value="📈 Quarterly Analysis (2+ changes per quarter)")
    q_title.font = subtitle_font
    row += 1
    
    if frequent_quarter_changes:
        # Headers
        headers = ["Quarter", "Module Name", "Change Count", "Changes Summary"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = thin_border
        row += 1
        
        # Group by quarter
        quarter_groups = {}
        for item in frequent_quarter_changes:
            quarter = item['period']
            if quarter not in quarter_groups:
                quarter_groups[quarter] = []
            quarter_groups[quarter].append(item)
        
        sorted_quarters = sorted(quarter_groups.keys(), key=get_quarter_sort_key)
        
        for quarter in sorted_quarters:
            for item in quarter_groups[quarter]:
                # Create summary text
                categories = {}
                for change in item['changes']:
                    cat = change['category']
                    categories[cat] = categories.get(cat, 0) + 1
                
                summary_parts = []
                for cat, count in categories.items():
                    summary_parts.append(f"{cat}: {count}")
                summary_text = "; ".join(summary_parts)
                
                ws.cell(row=row, column=1, value=quarter)
                ws.cell(row=row, column=2, value=item['module_name'])
                ws.cell(row=row, column=3, value=item['change_count'])
                summary_cell = ws.cell(row=row, column=4, value=summary_text)
                summary_cell.alignment = align_wrap
                
                # Apply borders
                for col in range(1, 5):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border
                row += 1
    else:
        ws.merge_cells(f'A{row}:D{row}')
        no_q_data = ws.cell(row=row, column=1, value="No modules found with 2+ changes per quarter")
        no_q_data.font = Font(italic=True, color="008000")
        row += 1
    
    row += 2
    
    # --- Yearly Changes Section ---
    ws.merge_cells(f'A{row}:D{row}')
    y_title = ws.cell(row=row, column=1, value="📊 Yearly Analysis (3+ changes per year)")
    y_title.font = subtitle_font
    row += 1
    
    if frequent_year_changes:
        # Headers
        headers = ["Year", "Module Name", "Change Count", "Changes Summary"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = thin_border
        row += 1
        
        # Group by year
        year_groups = {}
        for item in frequent_year_changes:
            year = item['period']
            if year not in year_groups:
                year_groups[year] = []
            year_groups[year].append(item)
        
        sorted_years = sorted(year_groups.keys(), reverse=True)
        
        for year in sorted_years:
            for item in year_groups[year]:
                # Create summary text
                categories = {}
                for change in item['changes']:
                    cat = change['category']
                    categories[cat] = categories.get(cat, 0) + 1
                
                summary_parts = []
                for cat, count in categories.items():
                    summary_parts.append(f"{cat}: {count}")
                summary_text = "; ".join(summary_parts)
                
                ws.cell(row=row, column=1, value=year)
                ws.cell(row=row, column=2, value=item['module_name'])
                ws.cell(row=row, column=3, value=item['change_count'])
                summary_cell = ws.cell(row=row, column=4, value=summary_text)
                summary_cell.alignment = align_wrap
                
                # Apply borders
                for col in range(1, 5):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border
                row += 1
    else:
        ws.merge_cells(f'A{row}:D{row}')
        no_y_data = ws.cell(row=row, column=1, value="No modules found with 3+ changes per year")
        no_y_data.font = Font(italic=True, color="008000")
        row += 1
    
    # --- Column widths ---
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 60
    
    print(f"Frequent changes worksheet created: {len(frequent_quarter_changes)} quarterly, {len(frequent_year_changes)} yearly") 